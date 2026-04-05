"""
Buat log sumber gambar dalam format Excel (.xlsx).
Scan otomatis semua file di folder sumber, generate MD5 hash, dan catat metadata.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import hashlib
import os
from datetime import datetime

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE = os.path.join(BASE_DIR, "log_sumber_gambar.xlsx")

# === DATA SUMBER LENGKAP (Opsi D naming) ===
# Format: (filename, folder, gender, age_group, sumber, url, platform, deskripsi, usia_estimasi)

DATA = [
    # =============================================
    # WANITA 25-39
    # =============================================
    ("W25_001.jpg", "contoh_wanita/25-39", "Wanita", "25-39",
     "Dr. JJ Wendel Plastic Surgery",
     "https://www.drjjwendel.com/wp-content/uploads/2022/07/before-botox_drjjwendel4.jpeg",
     "drjjwendel.com", "Before botox, front view, clinical, grey bg", "~26"),

    ("W25_002.jpg", "contoh_wanita/25-39", "Wanita", "25-39",
     "Dr. JJ Wendel Plastic Surgery",
     "https://www.drjjwendel.com/wp-content/uploads/2021/10/before-botox_drjjwendel2.jpeg",
     "drjjwendel.com", "Before botox, front view, blonde, blue eyes", "~30-35"),

    ("W25_003.jpg", "contoh_wanita/25-39", "Wanita", "25-39",
     "Dr. JJ Wendel Plastic Surgery",
     "https://www.drjjwendel.com/wp-content/uploads/2022/03/before-botox_drjjwendel4.jpeg",
     "drjjwendel.com", "Before botox, front view, blonde, natural", "~25"),

    ("W25_004.jpg", "contoh_wanita/25-39", "Wanita", "25-39",
     "Dr. JJ Wendel Plastic Surgery",
     "https://www.drjjwendel.com/wp-content/uploads/2020/12/before-botox-vollure_drjjwendel-3.jpg",
     "drjjwendel.com", "Before botox+vollure, front view, dark hair", "~35"),

    ("W25_005.jpg", "contoh_wanita/25-39", "Wanita", "25-39",
     "Dr. JJ Wendel Plastic Surgery",
     "https://www.drjjwendel.com/wp-content/uploads/2022/08/before-botox-11_drjjwendel.jpg",
     "drjjwendel.com", "Before botox, front view, pink hair", "~28"),

    ("W25_006.jpg", "contoh_wanita/25-39", "Wanita", "25-39",
     "ASPS - Brow Lift Gallery",
     "https://www1.plasticsurgery.org/include/images/photogallery/cases/5597/42219-116936a_scaled.jpg",
     "plasticsurgery.org", "Before brow lift, front view, clinical", "~38"),

    ("W25_007.jpg", "contoh_wanita/25-39", "Wanita", "25-39",
     "ASPS - Ear Surgery Gallery",
     "https://www1.plasticsurgery.org/include/images/photogallery/cases/10313/41364-115258a_scaled.jpg",
     "plasticsurgery.org", "Before ear surgery, front view", "~25"),

    ("W25_008.jpg", "contoh_wanita/25-39", "Wanita", "25-39",
     "ASPS - Rhinoplasty Gallery",
     "https://www1.plasticsurgery.org/include/images/photogallery/cases/105195/35464-102153a_scaled.jpg",
     "plasticsurgery.org", "Before rhinoplasty, front view", "~25"),

    ("W25_009.jpg", "contoh_wanita/25-39", "Wanita", "25-39",
     "ASPS - Dermal Fillers Gallery",
     "https://www1.plasticsurgery.org/include/images/photogallery/cases/112457/35568-102417a_scaled.jpg",
     "plasticsurgery.org", "Before filler, front view, headband pink", "~35-38"),

    ("W25_010.jpg", "contoh_wanita/25-39", "Wanita", "25-39",
     "Pexels", "https://www.pexels.com/photo/3760583/",
     "pexels.com", "Front view, orange bg, natural portrait", "~25"),

    ("W25_011.jpg", "contoh_wanita/25-39", "Wanita", "25-39",
     "ASPS - FL Dermal Fillers (Case 9701)",
     "https://www1.plasticsurgery.org/include/images/photogallery/cases/6499/9701-68941a.jpg",
     "plasticsurgery.org", "Before filler, front view, blonde, Florida", "~30-35"),

    ("W25_012.jpg", "contoh_wanita/25-39", "Wanita", "25-39",
     "ASPS - Naples FL Filler (Case 48635)",
     "https://www1.plasticsurgery.org/include/images/photogallery/cases/114944/48635-135039a_scaled.jpg",
     "plasticsurgery.org", "Before filler, front view, Naples FL", "~28-32"),

    # W25_013 - W25_021: Pexels API batch
    ("W25_013.jpg", "contoh_wanita/25-39", "Wanita", "25-39",
     "Pexels API", "https://www.pexels.com/photo/8244560/",
     "pexels.com", "Pexels API: woman headshot no makeup", "~25-35"),

    ("W25_014.jpg", "contoh_wanita/25-39", "Wanita", "25-39",
     "Pexels API", "https://www.pexels.com/photo/10648948/",
     "pexels.com", "Pexels API: woman headshot no makeup", "~25-35"),

    ("W25_015.jpg", "contoh_wanita/25-39", "Wanita", "25-39",
     "Pexels API", "https://www.pexels.com/photo/7298625/",
     "pexels.com", "Pexels API: woman headshot no makeup", "~25-35"),

    ("W25_016.jpg", "contoh_wanita/25-39", "Wanita", "25-39",
     "Pexels API", "https://www.pexels.com/photo/1552543/",
     "pexels.com", "Pexels API: woman headshot no makeup", "~25-35"),

    ("W25_017.jpg", "contoh_wanita/25-39", "Wanita", "25-39",
     "Pexels API", "https://www.pexels.com/photo/9421385/",
     "pexels.com", "Pexels API: woman headshot no makeup", "~25-35"),

    ("W25_018.jpg", "contoh_wanita/25-39", "Wanita", "25-39",
     "Pexels API", "https://www.pexels.com/photo/13771127/",
     "pexels.com", "Pexels API: woman headshot no makeup", "~25-35"),

    ("W25_019.jpg", "contoh_wanita/25-39", "Wanita", "25-39",
     "Pexels API", "https://www.pexels.com/photo/6659419/",
     "pexels.com", "Pexels API: woman headshot no makeup", "~25-35"),

    ("W25_020.jpg", "contoh_wanita/25-39", "Wanita", "25-39",
     "Pexels API", "https://www.pexels.com/photo/7994381/",
     "pexels.com", "Pexels API: woman headshot no makeup", "~25-35"),

    ("W25_021.jpg", "contoh_wanita/25-39", "Wanita", "25-39",
     "Pexels API", "https://www.pexels.com/photo/20068197/",
     "pexels.com", "Pexels API: woman headshot no makeup", "~25-35"),

    # =============================================
    # WANITA 40-65
    # =============================================
    ("W40_001.jpg", "contoh_wanita/40-65", "Wanita", "40-65",
     "Pexels API", "https://www.pexels.com/photo/6874469/",
     "pexels.com", "Mature caucasian woman face headshot natural aging", "~45-55"),

    ("W40_002.jpg", "contoh_wanita/40-65", "Wanita", "40-65",
     "Pexels API", "https://www.pexels.com/photo/12645006/",
     "pexels.com", "Mature caucasian woman headshot natural", "~45-55"),

    ("W40_003.jpg", "contoh_wanita/40-65", "Wanita", "40-65",
     "Pexels API", "https://www.pexels.com/photo/9969438/",
     "pexels.com", "Mature caucasian woman headshot natural", "~45-55"),

    ("W40_004.jpg", "contoh_wanita/40-65", "Wanita", "40-65",
     "Pexels API", "https://www.pexels.com/photo/6749928/",
     "pexels.com", "Mature caucasian woman headshot natural", "~45-55"),

    ("W40_005.jpg", "contoh_wanita/40-65", "Wanita", "40-65",
     "Pexels API", "https://www.pexels.com/photo/8727557/",
     "pexels.com", "Mature caucasian woman headshot natural", "~45-55"),

    ("W40_006.jpg", "contoh_wanita/40-65", "Wanita", "40-65",
     "Pexels API", "https://www.pexels.com/photo/8558835/",
     "pexels.com", "Mature caucasian woman headshot natural", "~45-55"),

    ("W40_007.jpg", "contoh_wanita/40-65", "Wanita", "40-65",
     "Pexels API", "https://www.pexels.com/photo/31233195/",
     "pexels.com", "Mature caucasian woman headshot natural", "~45-55"),

    ("W40_008.jpg", "contoh_wanita/40-65", "Wanita", "40-65",
     "Pexels API", "https://www.pexels.com/photo/5263308/",
     "pexels.com", "Mature caucasian woman headshot natural", "~40-50"),

    ("W40_009.jpg", "contoh_wanita/40-65", "Wanita", "40-65",
     "Pexels API", "https://www.pexels.com/photo/8727503/",
     "pexels.com", "Mature caucasian woman headshot natural", "~45-55"),

    ("W40_010.jpg", "contoh_wanita/40-65", "Wanita", "40-65",
     "Pexels API", "https://www.pexels.com/photo/5263318/",
     "pexels.com", "Mature caucasian woman headshot natural", "~40-50"),

    ("W40_011.jpg", "contoh_wanita/40-65", "Wanita", "40-65",
     "ASPS - Brow Lift Gallery",
     "https://www1.plasticsurgery.org/include/images/photogallery/cases/5597/42219-116936a_scaled.jpg",
     "plasticsurgery.org", "Before brow lift, front view, clinical", "~40-45"),

    # =============================================
    # PRIA 25-39
    # =============================================
    ("M25_001.jpg", "contoh_pria/25-39", "Pria", "25-39",
     "ASPS - Rhinoplasty (Case 31845)",
     "https://www1.plasticsurgery.org/include/images/photogallery/cases/111615/31845-93789a_scaled.jpg",
     "plasticsurgery.org", "Before rhinoplasty, front view, clinical", "~22-25"),

    ("M25_002.jpg", "contoh_pria/25-39", "Pria", "25-39",
     "ASPS - Rhinoplasty (Case 32844)",
     "https://www1.plasticsurgery.org/include/images/photogallery/cases/106234/32844-95988a_scaled.jpg",
     "plasticsurgery.org", "Before rhinoplasty, front view, clinical", "~20-25"),

    ("M25_003.jpg", "contoh_pria/25-39", "Pria", "25-39",
     "Pexels", "https://www.pexels.com/photo/30450838/",
     "pexels.com", "Young man headshot, front view, grey bg", "~22-25"),

    ("M25_004.jpg", "contoh_pria/25-39", "Pria", "25-39",
     "Pexels", "https://www.pexels.com/photo/614810/",
     "pexels.com", "Man close-up, front view, grey bg", "~25-28"),

    ("M25_005.jpg", "contoh_pria/25-39", "Pria", "25-39",
     "Pexels", "https://www.pexels.com/photo/3785077/",
     "pexels.com", "Man front view, grey bg, jacket", "~30-35"),

    ("M25_006.jpg", "contoh_pria/25-39", "Pria", "25-39",
     "Pexels", "https://www.pexels.com/photo/1933873/",
     "pexels.com", "Man front view, outdoor close-up", "~35-38"),

    ("M25_007.jpg", "contoh_pria/25-39", "Pria", "25-39",
     "Pexels", "https://www.pexels.com/photo/220453/",
     "pexels.com", "Man front view, grey bg, kacamata", "~25"),

    ("M25_008.jpg", "contoh_pria/25-39", "Pria", "25-39",
     "ASPS - Miami Beach Rhinoplasty (Case 44610)",
     "https://www1.plasticsurgery.org/include/images/photogallery/cases/114968/44610-123792a_scaled.jpg",
     "plasticsurgery.org", "Before rhinoplasty, front view, Miami Beach FL", "~35-38"),

    # M25_009 - M25_023: Pexels API batch
    ("M25_009.jpg", "contoh_pria/25-39", "Pria", "25-39",
     "Pexels API", "https://www.pexels.com/photo/10194766/",
     "pexels.com", "Pexels API: caucasian man face close up natural headshot", "~25-38"),

    ("M25_010.jpg", "contoh_pria/25-39", "Pria", "25-39",
     "Pexels API", "https://www.pexels.com/photo/7298876/",
     "pexels.com", "Pexels API: caucasian man face close up natural headshot", "~25-38"),

    ("M25_011.jpg", "contoh_pria/25-39", "Pria", "25-39",
     "Pexels API", "https://www.pexels.com/photo/15603012/",
     "pexels.com", "Pexels API: caucasian man face close up natural headshot", "~25-38"),

    ("M25_012.jpg", "contoh_pria/25-39", "Pria", "25-39",
     "Pexels API", "https://www.pexels.com/photo/5438476/",
     "pexels.com", "Pexels API: caucasian man face close up natural headshot", "~25-38"),

    ("M25_013.jpg", "contoh_pria/25-39", "Pria", "25-39",
     "Pexels API", "https://www.pexels.com/photo/10521294/",
     "pexels.com", "Pexels API: caucasian man face close up natural headshot", "~25-38"),

    ("M25_014.jpg", "contoh_pria/25-39", "Pria", "25-39",
     "Pexels API", "https://www.pexels.com/photo/17388022/",
     "pexels.com", "Pexels API: caucasian man face close up natural headshot", "~25-38"),

    ("M25_015.jpg", "contoh_pria/25-39", "Pria", "25-39",
     "Pexels API", "https://www.pexels.com/photo/8727383/",
     "pexels.com", "Pexels API: caucasian man face close up natural headshot", "~25-38"),

    ("M25_016.jpg", "contoh_pria/25-39", "Pria", "25-39",
     "Pexels API", "https://www.pexels.com/photo/8727558/",
     "pexels.com", "Pexels API: caucasian man face close up natural headshot", "~25-38"),

    ("M25_017.jpg", "contoh_pria/25-39", "Pria", "25-39",
     "Pexels API", "https://www.pexels.com/photo/8875609/",
     "pexels.com", "Pexels API: caucasian man face close up natural headshot", "~25-38"),

    ("M25_018.jpg", "contoh_pria/25-39", "Pria", "25-39",
     "Pexels API", "https://www.pexels.com/photo/18160382/",
     "pexels.com", "Pexels API: caucasian man face close up natural headshot", "~25-38"),

    ("M25_019.jpg", "contoh_pria/25-39", "Pria", "25-39",
     "Pexels API", "https://www.pexels.com/photo/34775253/",
     "pexels.com", "Pexels API: before dermal filler face man natural", "~25-38"),

    ("M25_020.jpg", "contoh_pria/25-39", "Pria", "25-39",
     "Pexels API", "https://www.pexels.com/photo/10194769/",
     "pexels.com", "Pexels API: before dermal filler face man natural", "~25-38"),

    ("M25_021.jpg", "contoh_pria/25-39", "Pria", "25-39",
     "Pexels API", "https://www.pexels.com/photo/6335086/",
     "pexels.com", "Pexels API: before dermal filler face man natural", "~25-38"),

    ("M25_022.jpg", "contoh_pria/25-39", "Pria", "25-39",
     "Pexels API", "https://www.pexels.com/photo/36613594/",
     "pexels.com", "Pexels API: before dermal filler face man natural", "~25-38"),

    ("M25_023.jpg", "contoh_pria/25-39", "Pria", "25-39",
     "Pexels API", "https://www.pexels.com/photo/6625959/",
     "pexels.com", "Pexels API: before dermal filler face man natural", "~25-38"),

    # =============================================
    # PRIA 40-65
    # =============================================
    ("M40_001.jpg", "contoh_pria/40-65", "Pria", "40-65",
     "ASPS - Facelift (Case 6765)",
     "https://www1.plasticsurgery.org/include/images/photogallery/cases/1673/6765-62166a_scaled.jpg",
     "plasticsurgery.org", "Before facelift, front view, clinical", "~40-45"),

    ("M40_002.jpg", "contoh_pria/40-65", "Pria", "40-65",
     "Pexels", "https://www.pexels.com/photo/2380794/",
     "pexels.com", "Man close-up, front view, botak+jenggot", "~40-45"),

    ("M40_003.jpg", "contoh_pria/40-65", "Pria", "40-65",
     "Pexels", "https://www.pexels.com/photo/2182970/",
     "pexels.com", "Man front view, kacamata, professional", "~40"),

    # M40_004 - M40_013: Pexels API batch (40-65 search)
    ("M40_004.jpg", "contoh_pria/40-65", "Pria", "40-65",
     "Pexels API", "https://www.pexels.com/photo/34253629/",
     "pexels.com", "Older man face portrait close up natural", "~45-55"),

    ("M40_005.jpg", "contoh_pria/40-65", "Pria", "40-65",
     "Pexels API", "https://www.pexels.com/photo/18030908/",
     "pexels.com", "Older man face portrait close up natural", "~45-55"),

    ("M40_006.jpg", "contoh_pria/40-65", "Pria", "40-65",
     "Pexels API", "https://www.pexels.com/photo/27640020/",
     "pexels.com", "Older man face portrait close up natural", "~45-55"),

    ("M40_007.jpg", "contoh_pria/40-65", "Pria", "40-65",
     "Pexels API", "https://www.pexels.com/photo/26842301/",
     "pexels.com", "Older man face portrait close up natural", "~45-55"),

    ("M40_008.jpg", "contoh_pria/40-65", "Pria", "40-65",
     "Pexels API", "https://www.pexels.com/photo/9830842/",
     "pexels.com", "Older man face portrait close up natural", "~45-55"),

    ("M40_009.jpg", "contoh_pria/40-65", "Pria", "40-65",
     "Pexels API", "https://www.pexels.com/photo/36263265/",
     "pexels.com", "Older man face portrait close up natural", "~45-55"),

    ("M40_010.jpg", "contoh_pria/40-65", "Pria", "40-65",
     "Pexels API", "https://www.pexels.com/photo/7298872/",
     "pexels.com", "Older man face portrait close up natural", "~50-60"),

    ("M40_011.jpg", "contoh_pria/40-65", "Pria", "40-65",
     "Pexels API", "https://www.pexels.com/photo/36305804/",
     "pexels.com", "Older man face portrait close up natural", "~45-55"),

    ("M40_012.jpg", "contoh_pria/40-65", "Pria", "40-65",
     "Pexels API", "https://www.pexels.com/photo/20449225/",
     "pexels.com", "Older man face portrait close up natural", "~45-55"),

    ("M40_013.jpg", "contoh_pria/40-65", "Pria", "40-65",
     "Pexels API", "https://www.pexels.com/photo/7298414/",
     "pexels.com", "Older man face portrait close up natural", "~50-60"),
]


def get_md5(filepath):
    if not os.path.exists(filepath):
        return "FILE NOT FOUND"
    with open(filepath, "rb") as f:
        return hashlib.md5(f.read()).hexdigest()


def create_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Log Sumber Gambar"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    w25_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    w40_fill = PatternFill(start_color="F8BBD0", end_color="F8BBD0", fill_type="solid")
    m25_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    m40_fill = PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    wrap = Alignment(wrap_text=True, vertical="top")

    # Headers
    headers = ["No", "Nama File", "Gender", "Kategori Usia", "Sumber",
               "URL / Link", "Platform", "Deskripsi", "Usia Estimasi",
               "MD5 Hash", "Tanggal Ditambahkan", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Data rows
    for i, (filename, folder, gender, age, sumber, url, platform, desc, usia) in enumerate(DATA, 1):
        row = i + 1
        filepath = os.path.join(BASE_DIR, folder, filename)
        md5 = get_md5(filepath)

        # Pilih warna berdasarkan kategori
        if filename.startswith("W25"):
            fill = w25_fill
        elif filename.startswith("W40"):
            fill = w40_fill
        elif filename.startswith("M25"):
            fill = m25_fill
        else:
            fill = m40_fill

        values = [i, filename, gender, age, sumber, url, platform,
                  desc, usia, md5, datetime.now().strftime("%Y-%m-%d"), "Aktif"]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill
            cell.border = border
            cell.alignment = wrap

    # Column widths
    widths = [5, 15, 10, 12, 35, 55, 18, 45, 15, 35, 15, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:L{len(DATA) + 1}"

    # === SHEET 2: Ringkasan ===
    ws2 = wb.create_sheet("Ringkasan")
    ws2["A1"] = "RINGKASAN LOG SUMBER GAMBAR"
    ws2["A1"].font = Font(bold=True, size=14)

    row = 3
    stats = [
        ("Total Gambar", len(DATA)),
        ("", ""),
        ("Wanita 25-39 (W25)", sum(1 for d in DATA if d[0].startswith("W25"))),
        ("Wanita 40-65 (W40)", sum(1 for d in DATA if d[0].startswith("W40"))),
        ("Pria 25-39 (M25)", sum(1 for d in DATA if d[0].startswith("M25"))),
        ("Pria 40-65 (M40)", sum(1 for d in DATA if d[0].startswith("M40"))),
        ("", ""),
        ("Platform:", "Jumlah"),
    ]

    for label, val in stats:
        ws2.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws2.cell(row=row, column=2, value=val)
        row += 1

    platforms = {}
    for d in DATA:
        p = d[6]
        platforms[p] = platforms.get(p, 0) + 1
    for platform, count in sorted(platforms.items()):
        ws2.cell(row=row, column=1, value=f"  {platform}")
        ws2.cell(row=row, column=2, value=count)
        row += 1

    row += 1
    ws2.cell(row=row, column=1, value="Terakhir Diperbarui").font = Font(bold=True)
    ws2.cell(row=row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M"))

    row += 2
    ws2.cell(row=row, column=1, value="CATATAN:").font = Font(bold=True, color="FF0000")
    for note in [
        "- Setiap gambar baru WAJIB dicek MD5 hash terhadap daftar ini",
        "- Tidak boleh ada wajah orang yang sama diekstrak 2x",
        "- Format nama file: W25_001 (Wanita 25-39), M40_001 (Man 40-65)",
        "- Referensikan file ini setiap kali menambah gambar baru",
    ]:
        row += 1
        ws2.cell(row=row, column=1, value=note)

    ws2.column_dimensions["A"].width = 55
    ws2.column_dimensions["B"].width = 20

    wb.save(OUTPUT_FILE)
    print(f"Excel disimpan: {OUTPUT_FILE}")
    print(f"Total: {len(DATA)} gambar")
    print(f"  W25: {sum(1 for d in DATA if d[0].startswith('W25'))}")
    print(f"  W40: {sum(1 for d in DATA if d[0].startswith('W40'))}")
    print(f"  M25: {sum(1 for d in DATA if d[0].startswith('M25'))}")
    print(f"  M40: {sum(1 for d in DATA if d[0].startswith('M40'))}")


if __name__ == "__main__":
    create_excel()
